import instaloader
import pandas as pd
from datetime import datetime
import locale

#turkce gunleri ve aylari duzeltmek icin
locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')
#instaloader dan ornek olustur
L=instaloader.Instaloader()
#kullanici ad
username=input("bir kullanici adini giriniz")
connt=int(input("kac adet gonderi analiz edilsin"))


#kullanici bilgilerini indir
profile =instaloader.Profile.from_username(L.context, username)
posts=profile.get_posts()

#analiz icin bir array
posts_data=[]
#kullanicinin gonderilerini al
for i ,post in enumerate(posts):
    if i>=connt:
        break
    post_info={
        'gun':post.date.strftime('%a'),
        'ay':post.date.strftime('%b'),
        'yil':post.date.year,
        'begeni sayisi':post.likes,
        'saat':post.date.strftime('%H%M%S'),
        'gonderi Linki':f'https//www.instagram.com/p/{post.shortcode}'
    }
    posts_data.append(post_info)

df=pd.DataFrame(posts_data)
#turkce karakter sorununu ortadan kaldirma
df['gun']=df['gun'].str.encode('utf8').str.decode('utf-8')
df['ay']=df['ay'].str.encode('utf8').str.decode('utf-8')

#excel dosyasina veri yazma
excel_file='insta_analiz.xlsx'
df.to_excel(excel_file, index=False,engine='openpyxl')

#excel dosyasu ac kosullu bicimlendir
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb=load_workbook(excel_file)
ws=wb.active

#begeni sayisi 2000 olan stunlari yesile boyayalim
for row in ws.iter_rows(min_row=2, max_row=4, min_col=len(df)+1, max_col=4):
    for cell in row:
        if cell.value>2000:
            cell.fill=PatternFill(start_color="00FF00",end_color="00FF00",fill_type="solid")

#excell dosyasini kaydet
wb.save(excel_file)














